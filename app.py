import streamlit as st
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
import shutil
from io import BytesIO
import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime


# ✅ 페이지 설정 (wide + 스타일 조정)
st.set_page_config(layout="wide")

st.markdown("""
    <style>
        .block-container {padding-top: 4rem;}
        .stButton > button:first-child {border: 2px solid #ffcccc;}
        td, th {text-align: center; vertical-align: middle;}
        th {background-color: #f5f5f5;}
        th:nth-child(1), td:nth-child(1),
        th:nth-child(2), td:nth-child(2),
        th:nth-child(3), td:nth-child(3) {text-align: left;}
        section.main > div {overflow-x: hidden;}
    </style>
""", unsafe_allow_html=True)

# ===========================================
# ✅ 함수: 데이터 수집
# ===========================================
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import pandas as pd
import yfinance as yf
import streamlit as st

# 날짜 계산
# def prepare_dates(selected_date):
#     기준일 = datetime.combine(selected_date, datetime.min.time()).replace(tzinfo=ZoneInfo("Asia/Seoul"))
#     start_date = 기준일 - timedelta(days=370)
    
#     # ✅ UTC 기준 시간 문제 방지: 오늘 날짜까지만 요청
#     # today_utc = datetime.utcnow().date()
#     # end_date = min(기준일 + timedelta(days=1), today_utc)
#     end_date = 기준일 + timedelta(days=1)
#     return 기준일, start_date, end_date

def prepare_dates(selected_date):
    기준일 = datetime.combine(selected_date, datetime.min.time()).replace(tzinfo=ZoneInfo("Asia/Seoul"))
    start_date = 기준일 - timedelta(days=370)

    # ✅ 오늘(UTC 기준)을 datetime 형식으로 변환
    today_utc = datetime.utcnow().replace(tzinfo=ZoneInfo("UTC"))
    today_utc_local = today_utc.astimezone(ZoneInfo("Asia/Seoul")).replace(hour=0, minute=0, second=0, microsecond=0)

    # ✅ 기준일 다음날과 오늘(로컬 기준) 중 작은 값으로 제한
    end_date = min(기준일 + timedelta(days=1), today_utc_local)

    return 기준일, start_date, end_date

# 영업일 추출
def get_business_days(start_date, end_date):
    all_dates = pd.date_range(start=start_date, end=end_date)
    return all_dates[all_dates.weekday < 5]

# 기준일 기준 기준연말, 전달말, 최근 5일 + 헤더 텍스트
def get_reference_dates(기준일, business_days):
    last_year = 기준일.year - 1
    last_year_end = max([d for d in business_days if d.year == last_year])

    prev_month = 기준일.month - 1 if 기준일.month > 1 else 12
    prev_year = 기준일.year if 기준일.month > 1 else 기준일.year - 1
    prev_month_end = datetime(prev_year, prev_month, 1, tzinfo=ZoneInfo("Asia/Seoul")) + timedelta(days=31)
    prev_month_end = prev_month_end.replace(day=1) - timedelta(days=1)
    prev_month_end = max([d for d in business_days if d.month == prev_month and d <= 기준일])

    recent_days = [d for d in business_days if d < 기준일][-5:]

    weekday_map = {0: "월", 1: "화", 2: "수", 3: "목", 4: "금", 5: "토", 6: "일"}
    headers = [
        f"{last_year % 100}년 未",
        f"{기준일.year % 100}.{prev_month}월",
    ] + [f"{d.month}/{d.day}({weekday_map[d.weekday()]})" for d in recent_days] + ["변동량", "변동률(%)"]

    return last_year_end, prev_month_end, recent_days, headers

# csv 로딩
def load_index_list(path):
    return pd.read_csv(path)

def fetch_from_naver(ticker_code, country, category, label):
    headers = {"User-Agent": "Mozilla/5.0"}

    world_codes = {"FX_USDVND", "FX_USDIDR", "FX_USDPHP", "FX_USDCNY"}
    if ticker_code in world_codes:
        url = f"https://finance.naver.com/marketindex/worldDailyQuote.naver?fdtc=4&marketindexCd={ticker_code}&page=1&pageSize=500"
    else:
        url = f"https://finance.naver.com/marketindex/exchangeDailyQuote.naver?marketindexCd={ticker_code}&page=1&pageSize=500"

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.select_one("table.tbl_exchange")
    rows = table.select("tbody > tr")

    records = []
    for row in rows:
        cols = row.select("td")
        if len(cols) < 2:
            continue
        try:
            raw_date = cols[0].text.strip()
            raw_close = cols[1].text.strip().replace(",", "")
            date_obj = datetime.strptime(raw_date, "%Y.%m.%d")
            close_val = float(raw_close)
            records.append([country, category, label, ticker_code, date_obj, close_val])
        except ValueError:
            continue

    df = pd.DataFrame(records, columns=["국가", "구분", "단위", "Ticker", "Date", "Close"])
    return df.sort_values("Date").reset_index(drop=True)

import yfinance as yf

def fetch_from_yfinance(ticker, start_date, end_date, country, category, label):
    try:
        data = yf.download(
            ticker,
            start=start_date.strftime("%Y-%m-%d"),
            end=end_date.strftime("%Y-%m-%d"),
            progress=False
        )
        # ✅ 데이터가 비어있는 경우 경고 출력
        if data.empty:
            st.warning(f"⚠️ {ticker} (yfinance): 조회된 데이터가 없습니다. 기간: {start_date.date()} ~ {end_date.date()}")
            return None

        if "Close" not in data.columns:
            st.warning(f"⚠️ {ticker} (yfinance): 'Close' 컬럼이 존재하지 않습니다.")
            return None

        data = data.reset_index()
        data = data[["Date", "Close"]]
        data["국가"] = country
        data["구분"] = category
        data["단위"] = label
        data["Ticker"] = ticker

        return data[["국가", "구분", "단위", "Ticker", "Date", "Close"]]

    except Exception as e:
        st.warning(f"{ticker} (yfinance) 에러: {e}")
        return None


# 개별 티커 데이터 다운로드 및 전처리
def fetch_price_data_by_source(ticker, start_date, end_date, row_meta):
    source = row_meta.get("수집출처", "yfinance")
    country = row_meta["국가"]
    category = row_meta["구분"]
    label = row_meta["항목명_짧은"]

    if source == "yfinance":
        return fetch_from_yfinance(ticker, start_date, end_date, country, category, label)
    elif source == "naver":
        return fetch_from_naver(ticker, country, category, label)
    else:
        st.warning(f"{ticker}의 수집출처 {source}는 지원되지 않습니다.")
        return None

# def download_and_clean_ticker_data(ticker, start_date, end_date, row_meta):
#     try:
#         data = yf.download(
#             ticker,
#             start=start_date.strftime("%Y-%m-%d"),
#             end=end_date.strftime("%Y-%m-%d"),
#             progress=False
#         )

#         if data.empty or 'Close' not in data.columns:
#             return None

#         data = data.reset_index()
#         data.columns = ["Date", "Close", "Open", "High", "Low", "Volume"]
#         data["국가"] = row_meta["국가"]
#         data["구분"] = row_meta["구분"]
#         data["단위"] = row_meta["항목명_짧은"]
#         data["Ticker"] = row_meta["티커"]

#         return data[["국가", "구분", "단위", "Ticker", "Date", "Close", "Open", "High", "Low", "Volume"]]

#     except Exception as e:
#         st.warning(f"{ticker} 에러: {e}")
#         return None

# 개별 레코드 계산
def calculate_price_record(price_series, row_meta, last_year_end, prev_month_end, recent_days, headers):
    record = {
        "국가": row_meta["국가"],
        "구분": row_meta["구분"],
        "단위": row_meta["항목명_짧은"],
    }

    # 인덱스를 날짜로 변환
    price_series.index = price_series.index.date

    values = {}
    for d, h in zip([last_year_end, prev_month_end] + recent_days, headers[:-2]):
        nearest_dates = [dt for dt in price_series.index if dt <= d.date()]
        if not nearest_dates:
            values[h] = None
        else:
            nearest_date = max(nearest_dates)
            value = price_series.loc[nearest_date]
            if isinstance(value, pd.Series):
                value = value.item()
            values[h] = value

    try:
        day_1 = values[headers[6]]
        day_2 = values[headers[5]]
        if (day_1 is None) or (day_2 is None) or (day_2 == 0):
            diff, rate = None, None
        else:
            diff = day_1 - day_2
            rate = (diff / day_2) * 100
    except:
        diff, rate = None, None

    values["변동량"] = diff
    values["변동률(%)"] = rate
    record.update(values)
    return record


# 전체 데이터 수집 및 조합
def collect_all_data(index_df, start_date, end_date, last_year_end, prev_month_end, recent_days, headers):
    records = []
    raw_records = []

    for _, row in index_df.iterrows():
        # data = download_and_clean_ticker_data(row["티커"], start_date, end_date, row)
        data = fetch_price_data_by_source(row["티커"], start_date, end_date, row)
        if data is None:
            continue

        data.columns = ["국가", "구분", "단위", "Ticker", "Date", "Close"]
        raw_records.append(data)
        price_series = data.set_index("Date")["Close"].dropna()
        record = calculate_price_record(price_series, row, last_year_end, prev_month_end, recent_days, headers)
        records.append(record)

    df_raw = pd.concat(raw_records, axis=0, ignore_index=True) if raw_records else pd.DataFrame()
    df_final = pd.DataFrame(records)
    return df_final, df_raw

# 결과 정렬
def sort_final_df(df_final):
    if not df_final.empty:
        country_order = {"베트남": 0, "인니": 1, "한국": 2, "필리핀": 3, "중국": 4, "미국": 5}
        category_order = {"주가": 0, "환율": 1}
        df_final["국가순서"] = df_final["국가"].map(country_order)
        df_final["구분순서"] = df_final["구분"].map(category_order)
        df_final = df_final.sort_values(["국가순서", "구분순서"]).drop(columns=["국가순서", "구분순서"]).reset_index(drop=True)
    return df_final

# ✅ 메인 함수
def fetch_data(selected_date):
    기준일, start_date, end_date = prepare_dates(selected_date)
    business_days = get_business_days(start_date, 기준일)
    last_year_end, prev_month_end, recent_days, headers = get_reference_dates(기준일, business_days)
    index_df = load_index_list("./index_list.csv")
    df_final, df_raw = collect_all_data(index_df, start_date, end_date, last_year_end, prev_month_end, recent_days, headers)
    df_final = sort_final_df(df_final)
    return df_final, df_raw, recent_days


# today = datetime.strptime('2025-04-30', '%Y-%m-%d').date()
# df_final, df_raw, recent_days = fetch_data(today)
# print(df_final)
# print(df_raw)

# ===========================================
# ✅ 메인 스트림릿 로직
# ===========================================

today_kst = datetime.now(ZoneInfo("Asia/Seoul")).date()

if "ready" not in st.session_state:
    st.session_state["ready"] = True
if "last_selected_date" not in st.session_state:
    st.session_state["last_selected_date"] = today_kst
if "downloading" not in st.session_state:
    st.session_state["downloading"] = False

col1, col2, col3, col4 = st.columns([2.5, 1, 1, 1])
with col1:
    st.title("국가별 주가 및 환율")
with col2:
    selected_date = st.date_input("\U0001F4C5 기준일을 선택하세요", today_kst, label_visibility="collapsed", max_value=today_kst)
with col3:
    fetch_button = st.button("\U0001F4E5 조회하기", use_container_width=True)
with col4:
    download_placeholder = st.empty()

if selected_date != st.session_state["last_selected_date"]:
    st.session_state["ready"] = False
    st.session_state["last_selected_date"] = selected_date

if 'df_final' not in st.session_state:
    with st.spinner("\u2728 초기 데이터 불러오는 중입니다..."):
        df_final, df_raw, recent_business_days = fetch_data(selected_date)
        st.session_state.df_final = df_final
        st.session_state.df_raw = df_raw
        st.session_state.recent_business_days = recent_business_days
        st.session_state["ready"] = True
elif fetch_button:
    with st.spinner("\ud83d\ude80 새 데이터 조회 중입니다... 조금만 기다려주세요!"):
        df_final, df_raw, recent_business_days = fetch_data(selected_date)
        st.session_state.df_final = df_final
        st.session_state.df_raw = df_raw
        st.session_state.recent_business_days = recent_business_days
        st.session_state["ready"] = True

# ✅ 테이블 표시
if 'df_final' in st.session_state and st.session_state["ready"] and not st.session_state["downloading"] and not st.session_state.df_final.empty:
    df_display = st.session_state.df_final.copy()

    def format_value(val, col=None):
        try:
            if pd.isna(val):
                return ""
            if isinstance(val, str):
                return val

            if col in ["변동량", "변동률(%)"]:
                val = round(val, 1)
                if val > 0:
                    return f"+{val:.1f}" if col == "변동량" else f"+{val:.1f}%"
                elif val < 0:
                    return f"△{abs(val):.1f}" if col == "변동량" else f"△{abs(val):.1f}%"
                else:
                    return f"{val:.1f}" if col == "변동량" else f"{val:.1f}%"

            if abs(val) >= 1000:
                return f"{int(round(val)):,}"
            else:
                return f"{val:.2f}"
        except:
            return val

    def style_row(row):
        styles = []
        for col in row.index:
            cell_style = "text-align: right;"
            if col in ["변동률(%)", "변동량"]:
                try:
                    val = row[col]
                    if pd.isna(val):
                        styles.append(cell_style)
                        continue
                    if val > 0:
                        cell_style += "color: blue;"
                    elif val < 0:
                        cell_style += "color: red;"
                except:
                    pass
            if col in ["국가", "구분", "단위"]:
                cell_style = "text-align: left;"
            styles.append(cell_style)
        return styles

    formatters = {col: (lambda val, c=col: format_value(val, c)) for col in df_display.columns}

    styled_table = (
        df_display.style
        .format(formatters, na_rep="")
        .apply(style_row, axis=1)
    )

    st.write(styled_table.to_html(index=False), unsafe_allow_html=True)

    # ✅ 엑셀 다운로드 버튼
    output = BytesIO()
    shutil.copy("./tmp/국가별 주가환율 테이블 템플릿_edit_v3.xlsx", "./tmp/temp_save.xlsx")
    wb = load_workbook("./tmp/temp_save.xlsx")

    ws_table = wb["table"]
    for row in ws_table.iter_rows(min_row=1, max_row=ws_table.max_row, min_col=1, max_col=ws_table.max_column):
        for cell in row:
            cell.value = None
    for idx, col_name in enumerate(st.session_state.df_final.columns, 1):
        ws_table.cell(row=1, column=idx).value = col_name
    for row_idx, row in enumerate(st.session_state.df_final.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_table.cell(row=row_idx, column=col_idx).value = value

    ws_name = wb["name"]
    start_day = st.session_state.recent_business_days[0]
    end_day = st.session_state.recent_business_days[-1]
    ws_name["A1"] = f"{start_day.month}/{start_day.day}~{end_day.month}/{end_day.day}"

    ws_raw = wb["rawdata"]
    for idx, col_name in enumerate(st.session_state.df_raw.columns, 1):
        ws_raw.cell(row=1, column=idx).value = col_name
        print(col_name)
    for row_idx, row in enumerate(st.session_state.df_raw.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws_raw.cell(row=row_idx, column=col_idx).value = value

    wb.save(output)
    output.seek(0)

    save_date = selected_date.strftime("%y%m%d")
    if download_placeholder.download_button(
        label="\U0001F4C4 엑셀 다운로드",
        data=output,
        file_name=f"국가별_주가환율정보_{save_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    ):
        st.session_state["downloading"] = True
    else:
        st.session_state["downloading"] = False